"""Extract raw data from data/results/ files (PDF, XLSX) to JSON.

Outputs to data/results_extracted/{univ_name}/{filename}.json

For each PDF page: extracts tables via find_tables() + page text.
For each XLSX sheet: extracts rows as table data.
Skips image-based PDFs and JPG files.
"""

import io
import json
import sys
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

RESULTS_DIR = Path("data/results")
OUTPUT_DIR = Path("data/results_extracted")


def extract_pdf(pdf_path: Path) -> dict | None:
    """Extract text and tables from a PDF. Returns None if image-based."""
    import fitz  # PyMuPDF

    doc = fitz.open(str(pdf_path))
    pages = []
    total_text = 0

    for page_num, page in enumerate(doc, 1):
        text = page.get_text()
        total_text += len(text)

        tables = []
        try:
            tab_finder = page.find_tables()
            for t in tab_finder.tables:
                rows = t.extract()
                if not rows:
                    continue
                tables.append({
                    "page": page_num,
                    "rows": len(rows),
                    "cols": t.col_count,
                    "data": rows,
                })
        except Exception:
            pass  # some pages fail table detection

        pages.append({
            "page_number": page_num,
            "text": text,
            "tables": tables,
        })

    doc.close()

    if total_text == 0:
        return None  # image-based PDF

    return {
        "source_file": pdf_path.name,
        "file_path": str(pdf_path),
        "format": "pdf",
        "pages": pages,
    }


def extract_xlsx(xlsx_path: Path) -> dict | None:
    """Extract sheets from an XLSX file."""
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert all cells to str or None
            cleaned = []
            for cell in row:
                if cell is None:
                    cleaned.append(None)
                else:
                    cleaned.append(str(cell).strip() if str(cell).strip() else None)
            # Skip completely empty rows
            if any(c is not None for c in cleaned):
                rows.append(cleaned)

        if rows:
            sheets.append({
                "sheet_name": sheet_name,
                "rows": len(rows),
                "cols": max(len(r) for r in rows) if rows else 0,
                "data": rows,
            })

    wb.close()

    if not sheets:
        return None

    return {
        "source_file": xlsx_path.name,
        "file_path": str(xlsx_path),
        "format": "xlsx",
        "sheets": sheets,
    }


def extract_xls(xls_path: Path) -> dict | None:
    """Extract sheets from an old-format XLS file using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(str(xls_path))
    sheets = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows = []
        for i in range(ws.nrows):
            cleaned = []
            for j in range(ws.ncols):
                val = ws.cell_value(i, j)
                if val is None or val == '':
                    cleaned.append(None)
                else:
                    s = str(val).strip()
                    # xlrd returns numbers as float; strip .0 for integers
                    if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
                        s = s[:-2]
                    cleaned.append(s if s else None)
            if any(c is not None for c in cleaned):
                rows.append(cleaned)

        if rows:
            sheets.append({
                "sheet_name": sheet_name,
                "rows": len(rows),
                "cols": max(len(r) for r in rows) if rows else 0,
                "data": rows,
            })

    if not sheets:
        return None

    return {
        "source_file": xls_path.name,
        "file_path": str(xls_path),
        "format": "xlsx",  # same format tag so batch extractor handles it
        "sheets": sheets,
    }


def process_university(univ_dir: Path, output_univ_dir: Path) -> dict:
    """Process all extractable files in a university directory."""
    stats = {"pdfs": 0, "xlsxs": 0, "skipped": 0, "errors": 0}

    for f in sorted(univ_dir.iterdir()):
        suffix = f.suffix.lower()
        out_path = output_univ_dir / (f.stem + ".json")

        # Skip already extracted files
        if out_path.exists():
            continue

        data = None
        try:
            if suffix == ".pdf":
                data = extract_pdf(f)
                if data:
                    stats["pdfs"] += 1
                else:
                    stats["skipped"] += 1
                    continue
            elif suffix == ".xlsx":
                data = extract_xlsx(f)
                if data:
                    stats["xlsxs"] += 1
                else:
                    stats["skipped"] += 1
                    continue
            elif suffix == ".xls":
                data = extract_xls(f)
                if data:
                    stats["xlsxs"] += 1
                else:
                    stats["skipped"] += 1
                    continue
            else:
                stats["skipped"] += 1
                continue
        except Exception as e:
            print(f"  ERROR {f.name}: {e}")
            stats["errors"] += 1
            continue

        output_univ_dir.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as fp:
            json.dump(data, fp, ensure_ascii=False, indent=2)

    return stats


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Extract result files to JSON")
    parser.add_argument("--univ", help="Process only this university (substring match)")
    parser.add_argument("--force", action="store_true", help="Re-extract already done files")
    args = parser.parse_args()

    univ_dirs = sorted(RESULTS_DIR.iterdir())
    if args.univ:
        univ_dirs = [d for d in univ_dirs if args.univ in d.name]

    total = {"pdfs": 0, "xlsxs": 0, "skipped": 0, "errors": 0}
    done = 0

    for univ_dir in univ_dirs:
        if not univ_dir.is_dir():
            continue

        output_univ_dir = OUTPUT_DIR / univ_dir.name

        if args.force:
            # Remove existing JSONs to force re-extraction
            for f in output_univ_dir.glob("*.json"):
                f.unlink()

        stats = process_university(univ_dir, output_univ_dir)
        if stats["pdfs"] + stats["xlsxs"] + stats["errors"] > 0:
            print(f"{univ_dir.name}: +{stats['pdfs']}pdf +{stats['xlsxs']}xlsx "
                  f"skip={stats['skipped']} err={stats['errors']}")

        for k in total:
            total[k] += stats[k]
        done += 1

    print(f"\nDone: {done} universities")
    print(f"Extracted: {total['pdfs']} PDFs, {total['xlsxs']} XLSXs")
    print(f"Skipped (image/other): {total['skipped']}, Errors: {total['errors']}")

    # List output JSON files
    out_files = list(OUTPUT_DIR.rglob("*.json"))
    print(f"Total JSON files in {OUTPUT_DIR}: {len(out_files)}")


if __name__ == "__main__":
    main()
